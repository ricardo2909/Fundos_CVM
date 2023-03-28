import streamlit as st
import pandas as pd
import requests
import zipfile
import datetime as dt
from openpyxl import load_workbook
from datetime import datetime
import os
import pickle
import time


# Definindo a função para consultar os fundos
@st.cache_data()
def consultar_fundos(cnpjs,data):
    
    ano = data[:4]
    mes = data[4:6]
    url = f"https://dados.cvm.gov.br/dados/FI/DOC/INF_DIARIO/DADOS/inf_diario_fi_{ano}{mes}.zip"

    download = requests.get(url)

    with open(f"inf_diario_fi_{ano}{mes}.zip", "wb") as arquivo_cvm:
        arquivo_cvm.write(download.content)

    arquivo_zip = zipfile.ZipFile(f"inf_diario_fi_{ano}{mes}.zip")

    dados_fundos = pd.read_csv(arquivo_zip.open(arquivo_zip.namelist()[0]), sep=";", encoding="ISO-8859-1", dtype= str)

    dados_cadastro = pd.read_csv('https://dados.cvm.gov.br/dados/FI/CAD/DADOS/cad_fi.csv', sep=";", encoding="ISO-8859-1", dtype={'CNPJ_FUNDO': str})

    dados_cadastro = dados_cadastro[['CNPJ_FUNDO', 'DENOM_SOCIAL']]
    dados_cadastro = dados_cadastro.drop_duplicates()

    base_final = pd.merge(dados_fundos, dados_cadastro, how='left', left_on='CNPJ_FUNDO', right_on='CNPJ_FUNDO')

    base_final = base_final[['DT_COMPTC', 'CNPJ_FUNDO', 'DENOM_SOCIAL', 'VL_QUOTA', 'VL_PATRIM_LIQ']]

    dia = int(data[6:8])

    base_filtro = base_final[base_final['CNPJ_FUNDO'].isin(cnpjs)]
    base_filtro = base_filtro[base_filtro['DT_COMPTC'] == f"{ano}-{mes}-{dia}"]
    base_filtro = base_filtro[['DT_COMPTC','CNPJ_FUNDO', 'DENOM_SOCIAL', 'VL_QUOTA']]
    base_filtro['VL_QUOTA'] = base_filtro['VL_QUOTA'].str.replace('.', ',')
    return base_filtro

def app():
    # Lendo a lista de CNPJs salvos do arquivo texto
    if os.path.isfile("cnpjs.txt"):
        with open("cnpjs.txt", "r") as arquivo:
            cnpjs_salvos = arquivo.read().splitlines()
    else:
        cnpjs_salvos = []

    st.title("Consulta de Fundos")

    # Criando as caixas de texto para receber os CNPJs e a data
    cnpj_input = st.text_input("Digite o CNPJ do fundo")

    # Botão para adicionar o CNPJ digitado à lista de CNPJs salvos
    if st.button("Adicionar CNPJ"):
        if cnpj_input:
            cnpj_list = [cnpj.strip() for cnpj in cnpj_input.split(",")]  # dividir entrada do usuário em uma lista de CNPJs
            cnpjs_validos = []  # criar lista para os CNPJs válidos
            for cnpj in cnpj_list:
                if cnpj not in cnpjs_salvos:
                    # Verificar CNPJ e adicionar à lista
                    cnpjs_salvos.append(cnpj)
                    msg = st.empty()  # criar espaço vazio na interface do usuário
                    msg.success(f"CNPJ {cnpj} adicionado com sucesso!")
                    time.sleep(1)  # aguardar 3 segundos
                    msg.empty()  # limpar mensagem após 3 segundos
                else:
                    msg = st.empty()
                    msg.warning(f"CNPJ {cnpj} inválido ou já existe na lista.")
                    time.sleep(1)
                    msg.empty()

            # Salvando a lista de CNPJs no arquivo texto
            with open("cnpjs.txt", "w") as arquivo:
                arquivo.write("\n".join(cnpjs_salvos))
        else:
            msg = st.empty()
            msg.warning("Digite um CNPJ.")
            time.sleep(1)
            msg.empty()
        

    # Caixa de seleção para remover um CNPJ da lista
    cnpj_remover = st.selectbox("Remover CNPJ", [None, *cnpjs_salvos])

    # Botão para remover o CNPJ selecionado da lista
    if st.button("Remover CNPJ"):
        if cnpj_remover:
            cnpjs_salvos.remove(cnpj_remover)
            st.success("CNPJ removido com sucesso!")
            # Salvando a lista de CNPJs no arquivo texto
            with open("cnpjs.txt", "w") as arquivo:
                arquivo.write("\n".join(cnpjs_salvos))
        else:
            st.warning("Selecione um CNPJ para remover.")

    # Botão para exibir/ocultar a tabela com a lista de CNPJs salvos
    exibir_tabela = st.checkbox("Mostrar lista de CNPJs salvos")
    if exibir_tabela:
        # Exibindo a lista de CNPJs salvos
        if cnpjs_salvos:
            st.write("CNPJs salvos:")
            for cnpj in cnpjs_salvos:
                st.write(f"- {cnpj}")
            st.write("\n")
        else:
            st.warning("Não há CNPJs salvos.")
    #pegar a data de hoje - 2 e transformar no padrao aceito pela API
    data_padrao =  dt.date.today() - dt.timedelta(days=2)
    
    # Se houver CNPJs salvos, exibe o campo para selecionar a data e o botão para consulta
    if cnpjs_salvos:
        data = st.date_input("Selecione a data que deseja consultar", value=data_padrao, max_value=dt.date.today())

        # Formatando a data para o formato AAAAMMDD
        data_str = data.strftime('%Y%m%d')

        # Chamando a função para consultar os fundos
        resultado = consultar_fundos(cnpjs_salvos, data_str)

        # Exibindo o resultado na tela
        if resultado.empty:
            st.warning("Não foram encontrados resultados para os CNPJs e a data informados.")
        else:
            st.write(resultado)
    else:
        st.warning("Digite um CNPJ e clique em 'Adicionar CNPJ' para começar a consultar.")

    nome_arquivo = st.text_input("Digite o nome do arquivo para salvar (sem extensão)", value = f"Fundos_{data_str}", max_chars=50)
    # Botão para exportar a tabela resultante em diferentes formatos
    formato_exportacao = st.selectbox("Selecione o formato de exportação", ["Excel", "CSV", "PDF"])
    if st.button("Exportar tabela"):
        if resultado.empty:
            st.warning("Não há tabela para exportar.")
        else:
            # Definindo o nome do arquivo de acordo com o que foi digitado pelo usuário
            if nome_arquivo:
                nome_arquivo = nome_arquivo.replace(" ", "_")
            else:
                nome_arquivo = "resultado"
            if formato_exportacao == "Excel":
                excel_file = resultado.to_excel(f"{nome_arquivo}.xlsx", index=False)
            elif formato_exportacao == "CSV":
                csv_file = resultado.to_csv(f"{nome_arquivo}.csv", index=False)
            elif formato_exportacao == "PDF":
                pdf_file = resultado.to_pdf(f"{nome_arquivo}.pdf", index=False)
            st.success(f"Tabela exportada como {nome_arquivo}.{formato_exportacao}")





if __name__ == "__main__":
    app()







#salvar num arquivo excel na aba do dia
# nome_da_aba = f"{dia}_{mes}_{ano}"
# nome_arquivo = "Fundos_CVM.xlsx"
# book = load_workbook(nome_arquivo)
# if nome_da_aba in book.sheetnames:
#     # se a aba já existe, sobrescreve os dados
#     writer = pd.ExcelWriter(nome_arquivo, engine='openpyxl')
#     writer.book = book
#     writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#     base_filtro.to_excel(writer, sheet_name=nome_da_aba, index=False)
#     writer.save()
# else:
#     # se a aba não existe, cria uma nova aba
#     with pd.ExcelWriter(nome_arquivo, engine='openpyxl') as writer:
#         writer.book = book
#         base_filtro.to_excel(writer, sheet_name=nome_da_aba, index=False)
#         writer.save()

